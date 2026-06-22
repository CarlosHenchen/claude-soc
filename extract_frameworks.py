# -*- coding: utf-8 -*-
"""Extracts frameworks into assets/frameworks.json with SOC-CMM formula fidelity.

Output schema (single JSON object):
{
  "viaconnect": <framework>,
  "soc": { "domains":[ { key,name, aspects:[ { code,name,num,hasCapability,
              questions:[ {code,text,type:'M'|'C',importance,scaleMax,
                           levels:[{v,label}], remark} ] } ] } ] }
}
The app builds 3 frameworks: viaconnect, soc-cmm-basic, soc-cmm-advanced
(the last two share `soc` but differ in importance/target editing).

SOC-CMM scoring (transcribed from the workbook formulas):
  factor(importance) = {1:0, 2:.5, 3:1, 4:2, 5:4}   (_Score matrix 4)
  per element: I = answer*factor ; H = factor ; J = 5*factor
  aspect %    : K = max(0, 100*(ΣI-ΣH)/(ΣJ-ΣH))     (Maturity uses type M, Capability type C)
  Maturity radar = 5*K/100 ; Capability radar = 3*K/100
"""
import openpyxl, warnings, re, json, os, sys, importlib.util, types
warnings.filterwarnings("ignore")
from openpyxl.utils import get_column_letter as L

HERE = os.path.dirname(os.path.abspath(__file__))
# Source spreadsheets live in the repo's `frameworks/` folder so the build is
# reproducible on any machine (Linux/CI included). Override with FRAMEWORKS_DIR.
DL = (os.environ.get("FRAMEWORKS_DIR") or os.path.join(HERE, "frameworks")).rstrip("/\\") + "/"
DOMS = {"Business", "People", "Process", "Technology", "Services"}
DOMORDER = ["Business", "People", "Process", "Technology", "Services"]
ELEM = re.compile(r'^\d+(\.\d+){1,2}$')
GELEM = re.compile(r'^([A-Z]{1,3})\s+(\d+(?:\.\d+){1,2})$')
GASPECT = re.compile(r'^([A-Z]{1,3})(\d+)\s*-\s*(.+)$')
GENERIC_M = {0: "Inexistente", 1: "Inicial", 2: "Limitado", 3: "Definido", 4: "Gerenciado", 5: "Otimizado"}
GENERIC_C = {0: "Não / inexistente", 1: "Incompleto", 2: "Parcial", 3: "Médio", 4: "Quase completo", 5: "Completo"}

# ---- Viaconnect (reuse existing modular parser) -----------------------------
pkg = types.ModuleType("ing"); sys.modules["ing"] = pkg
def _load(name, path, p=None):
    spec = importlib.util.spec_from_file_location(name, path)
    m = importlib.util.module_from_spec(spec)
    if p: m.__package__ = p
    sys.modules[name] = m; spec.loader.exec_module(m); return m
_load("ing.base", os.path.join(HERE, "ingest/base.py"))
vp = _load("ing.viaconnect_parser", os.path.join(HERE, "ingest/viaconnect_parser.py"), p="ing")

def _scale_blocks(path):
    """Parse section 3 (Escalas de maturidade) of the Framework sheet into blocks
    with the EXACT level name + description for each scale."""
    ws = openpyxl.load_workbook(path, data_only=True)["Framework"]
    blocks = {}
    cur = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        atxt = str(a).strip() if a is not None else ""
        if atxt.lower().startswith("escala"):
            cur = atxt
            blocks[cur] = {}
        elif cur is not None and atxt.isdigit() and b not in (None, ""):
            blocks[cur][int(atxt)] = {"name": str(b).strip(), "desc": (str(c).strip() if c not in (None, "") else "")}
        elif atxt.startswith("4."):
            cur = None
    return blocks


# domain.key -> keyword identifying its scale block in section 3
_SCALE_KW = {
    "CTI": "cti-cmm", "Threat Hunting": "hmm",
    "Detection Eng": "1–5", "DFIR": "1–5", "Vuln Mgmt": "1–5",
    "CSIRT": "sim3", "PSIRT": "psirt",
    "Deception": "0–5", "Automacao": "0–5",
}


def viaconnect():
    path = os.path.join(HERE, "frameworks/Viaconnect_HealthCheck_Dominios_SOC_COMPLETOv2.xlsx")
    fw = vp.ViaconnectParser().parse(path)
    blocks = _scale_blocks(path)

    def block_for(domain_key):
        kw = _SCALE_KW.get(domain_key, "")
        for title, lv in blocks.items():
            if kw and kw in title.lower():
                return lv
        return {}

    out = {"slug": "viaconnect-soc", "name": "Domínios (Viaconnect Health Check)",
           "kind": "viaconnect", "scaleNote": "Escala própria por domínio", "domains": []}
    for d in fw.domains:
        sb = block_for(d.key)
        # Levels use the EXACT name + description from the spreadsheet's scale section,
        # falling back to the dropdown label when a description is unavailable.
        levels = []
        for o in d.scale.options:
            info = sb.get(o.value)
            levels.append({
                "v": o.value,
                "label": (info["name"] if info else o.label),
                "desc": (info["desc"] if info else ""),
            })
        dom = {"key": d.key, "name": d.name, "aspects": []}
        for c in d.controls:
            asp = {"code": c.code, "name": c.name, "num": 0, "hasCapability": False, "questions": []}
            for q in c.questions:
                asp["questions"].append({"code": q.code, "text": q.text, "type": "M",
                    "importance": 3, "scaleMax": d.scale.max_value, "levels": levels,
                    "remark": q.guidance or ""})
            dom["aspects"].append(asp)
        out["domains"].append(dom)
    return out

# ---- SOC-CMM (formula-faithful) --------------------------------------------
def soc_catalog(path):
    F = openpyxl.load_workbook(path, data_only=False)
    # guidance: code -> {v: text}
    g = F["_Guidance"]; gmap = {}; cur = None
    for r in range(1, g.max_row + 1):
        a = g.cell(r, 1).value; b = g.cell(r, 2).value; c = g.cell(r, 3).value
        a = str(a).strip() if a is not None else ""
        if GELEM.match(a):
            cur = re.sub(r'\s+', ' ', a); gmap.setdefault(cur, {})
        elif a:
            cur = None
        if cur and b is not None and str(b).strip() != "":
            try: gmap[cur][int(b)] = (str(c).strip() if c else "")
            except ValueError: pass
    # aspect sheets: text + remark by code
    textmap, remarkmap = {}, {}
    for sh in F.sheetnames:
        if " - " not in sh: continue
        dom = sh.split(" - ")[0]
        if dom not in DOMS: continue
        letter = {"Business": "B", "People": "P", "Process": "M", "Technology": "T", "Services": "S"}[dom]
        ws = F[sh]
        rc = None
        for r in range(1, 14):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() == "remarks": rc = c
            if rc: break
        for r in range(1, ws.max_row + 1):
            b = ws.cell(r, 2).value; c = ws.cell(r, 3).value
            b = str(b).strip() if b is not None else ""
            if ELEM.match(b) and c not in (None, "", "None"):
                code = f"{letter} {b}"; textmap[code] = str(c).strip()
                if rc:
                    rv = ws.cell(r, rc).value
                    if rv not in (None, "", "None"): remarkmap[code] = str(rv).strip()
    # Results-OVR: ordered aspects (domain, name, mrow, crow)
    ov = F["Results - OVR"]; aspects = []; dom = None
    for r in range(10, ov.max_row + 1):
        b = ov.cell(r, 2).value; c = ov.cell(r, 3).value; d = ov.cell(r, 4).value
        h = ov.cell(r, 8).value; l = ov.cell(r, 12).value
        if isinstance(b, str) and b.strip() in DOMS: dom = b.strip()
        if c == "overall" or not d: continue
        if not (isinstance(h, str) and "_Output!K" in h): continue
        mk = re.search(r"_Output!K(\d+)", h)
        ck = re.search(r"_Output!K(\d+)", l) if isinstance(l, str) else None
        aspects.append({"domain": dom, "raw": str(d).strip(),
                        "mrow": int(mk.group(1)) if mk else None,
                        "crow": int(ck.group(1)) if ck else None})
    # _Output: elements (code,type,importance) + sum-row patterns
    out = F["_Output"]
    elems = []
    for r in range(1, out.max_row + 1):
        a = out.cell(r, 1).value; t = out.cell(r, 3).value; e = out.cell(r, 5).value
        if isinstance(a, str) and re.match(r'^[A-Z]{1,3}\s+\d', a) and t in ("M", "C"):
            elems.append({"code": re.sub(r'\s+', ' ', a.strip()), "type": t,
                          "imp": int(e) if isinstance(e, (int, float)) else 3})
    def pattern(row):
        f = out.cell(row, 4).value
        m = re.search(r'\$A:\$A,"([^"]+)"', f) if isinstance(f, str) else None
        return m.group(1).replace("*", "") if m else None

    def text_for(code):
        if code in textmap: return textmap[code]
        parts = code.split(" "); seg = parts[1].split(".")
        if len(seg) > 1:
            p = f"{parts[0]} {'.'.join(seg[:-1])}"
            if p in textmap: return textmap[p]
        return code

    def build_levels(code, typ):
        base = GENERIC_M if typ == "M" else GENERIC_C
        lv = gmap.get(code, {})
        out = []
        for v in range(0, 6):
            txt = lv.get(v, "").strip()
            out.append({"v": v, "label": (base[v] + " — " + txt) if txt and txt != base[v] else base[v]})
        return out

    def mk_questions(base, typ):
        qs = []
        for x in elems:
            if x["type"] == typ and x["code"].startswith(base):
                qs.append({"code": x["code"], "text": text_for(x["code"]), "type": typ,
                           "importance": x["imp"] or 3, "scaleMax": 5,
                           "levels": build_levels(x["code"], typ),
                           "remark": remarkmap.get(x["code"], "")})
        return qs

    domains = {}
    for asp in aspects:
        base = pattern(asp["mrow"]) if asp["mrow"] else None
        if not base: continue
        name = re.sub(r'^\d+\.\s*', '', asp["raw"])
        num = int(re.match(r'^(\d+)\.', asp["raw"]).group(1)) if re.match(r'^(\d+)\.', asp["raw"]) else 0
        code = base.strip()
        Q = mk_questions(base, "M") + (mk_questions(base, "C") if asp["crow"] else [])
        a = {"code": code, "name": name, "num": num, "hasCapability": bool(asp["crow"]), "questions": Q}
        domains.setdefault(asp["domain"], []).append(a)
    return {"domains": [{"key": d, "name": d, "aspects": domains[d]} for d in DOMORDER if d in domains]}

# ----------------------------------------------------------------------------
result = {"viaconnect": viaconnect(),
          "soc": soc_catalog(DL + "63-soc-cmm-242-advanced.xlsx")}
with open(os.path.join(HERE, "assets", "frameworks.json"), "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False)

soc = result["soc"]
nm = sum(len([q for q in a["questions"] if q["type"] == "M"]) for d in soc["domains"] for a in d["aspects"])
nc = sum(len([q for q in a["questions"] if q["type"] == "C"]) for d in soc["domains"] for a in d["aspects"])
na = sum(len(d["aspects"]) for d in soc["domains"])
print(f"viaconnect domains={len(result['viaconnect']['domains'])}")
print(f"soc domains={len(soc['domains'])} aspects={na} M={nm} C={nc}")
for d in soc["domains"]:
    print(f"  {d['key']:11}", ", ".join(f"{a['name']}({len([q for q in a['questions'] if q['type']=='M'])}M"
          + (f"+{len([q for q in a['questions'] if q['type']=='C'])}C" if a['hasCapability'] else "") + ")" for a in d["aspects"]))
print("bytes:", os.path.getsize(os.path.join(HERE, "assets", "frameworks.json")))
