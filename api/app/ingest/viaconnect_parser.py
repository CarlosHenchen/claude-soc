"""Parser for the Viaconnect Health Check SOC workbook.

Layout (validated against the real file)
-----------------------------------------
* Sheet "Framework": cover + reference-models table (rows 10-18: Aba | Modelo |
  Mantenedor | Escala | Referência) + scale descriptions. Used for domain metadata.
* Sheet "Dash": consolidated dashboard (computed in-app, not parsed).
* 9 domain sheets (one per assessment area). Each has:
    - row 1, col A: full domain title.
    - row 4: header row -> ID | Item de avaliação | Orientação para o avaliador |
      Avaliação atual | Nível | Meta | Nível meta | Gap | Evidências | ... | side table
    - rows 5+: either a SUBDOMAIN header (col A text, col B empty) or an ITEM
      (col A = "CTI-01", col B = item text, col C = guidance).
    - a single list data-validation per sheet defines the maturity scale/dropdown.

If the real layout changes, adjust the constants / `_is_item` below; the rest of
the pipeline (seeder, scoring, API) is decoupled from these details.
"""
from __future__ import annotations

import hashlib
import re

import openpyxl

from .base import (
    ParsedControl,
    ParsedDomain,
    ParsedFramework,
    ParsedQuestion,
    ParsedScale,
    ParsedScaleOption,
)

# Worksheets that are NOT assessment domains.
NON_DOMAIN_SHEETS = {"Framework", "Dash"}

# Order of the 9 domains as they appear in the workbook / reference table.
DOMAIN_ORDER = [
    "CTI", "Threat Hunting", "Detection Eng", "DFIR", "CSIRT",
    "PSIRT", "Deception", "Vuln Mgmt", "Automacao",
]

# Data rows start at row 5 (row 4 is the header).
DATA_START_ROW = 5
COL_ID = 1          # A
COL_ITEM = 2        # B
COL_GUIDANCE = 3    # C

# An assessment item id like CTI-01, TH-05, IR-12, CS-45, AU-09.
_ITEM_RE = re.compile(r"^[A-Z]{2,4}-\d+", re.IGNORECASE)
# Separators used inside subdomain headers ("PROGRAM · Gestão...").
_SEP_RE = re.compile(r"\s[·—–-]\s")
# Group-header prefixes whose meaningful name is on the RIGHT of the separator.
_PREFIX_RE = re.compile(r"^(Pilar|Fase|Área|Area)\b", re.IGNORECASE)


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _is_item(col_a: str, col_b: str) -> bool:
    return bool(_ITEM_RE.match(col_a.strip()))


def _split_header(header: str) -> tuple[str, str]:
    """Return (code, full_name) for a subdomain header row."""
    full = header.strip()
    parts = _SEP_RE.split(full, maxsplit=1)
    if len(parts) == 2:
        left, right = parts[0].strip(), parts[1].strip()
        code = right if _PREFIX_RE.match(left) else left
    else:
        code = full
    return code[:120], full[:500]


def _parse_scale(ws) -> ParsedScale:
    """Build the maturity scale from the worksheet's list data-validation."""
    options: list[ParsedScaleOption] = []
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue
        raw = dv.formula1.strip().strip('"')
        for token in raw.split(","):
            token = token.strip()
            if not token:
                continue
            m = re.match(r"^(\d+)\s*-\s*(.+)$", token)
            if m:
                options.append(ParsedScaleOption(value=int(m.group(1)), label=m.group(2).strip()))
            else:
                options.append(ParsedScaleOption(value=len(options), label=token))
        break  # one list validation per sheet

    if not options:  # defensive fallback
        options = [ParsedScaleOption(value=i, label=str(i)) for i in range(6)]

    values = [o.value for o in options]
    lo, hi = min(values), max(values)
    # Deterministic key so identical scales (e.g. CMMI 1-5) dedupe across domains.
    digest = hashlib.md5("|".join(f"{o.value}:{o.label}" for o in options).encode()).hexdigest()[:8]
    return ParsedScale(
        key=f"scl_{lo}_{hi}_{digest}",
        name=f"Escala {lo}–{hi}",
        min_value=lo,
        max_value=hi,
        options=options,
    )


def _read_reference_table(wb) -> dict[str, dict]:
    """rows 10-18 of the Framework sheet -> {aba: {model, maintainer, scale, url}}."""
    ws = wb["Framework"]
    refs: dict[str, dict] = {}
    for r in range(10, ws.max_row + 1):
        aba = _clean(ws.cell(r, 1).value)
        if not aba or aba.lower() == "aba":
            continue
        refs[aba] = {
            "reference_model": _clean(ws.cell(r, 2).value) or None,
            "maintainer": _clean(ws.cell(r, 3).value) or None,
            "scale_text": _clean(ws.cell(r, 4).value) or None,
            "reference_url": _clean(ws.cell(r, 5).value) or None,
        }
    return refs


class ViaconnectParser:
    """Parses the Viaconnect Health Check workbook into a ParsedFramework."""

    slug = "viaconnect-soc"
    name = "Viaconnect Health Check SOC"

    def parse(self, path: str) -> ParsedFramework:
        wb = openpyxl.load_workbook(path, data_only=True)
        refs = _read_reference_table(wb)

        framework = ParsedFramework(
            slug=self.slug,
            name=self.name,
            description=_clean(wb["Framework"].cell(1, 1).value) or self.name,
            source_file=path.replace("\\", "/").split("/")[-1],
        )

        domain_sheets = [s for s in wb.sheetnames if s not in NON_DOMAIN_SHEETS]
        ordered = [d for d in DOMAIN_ORDER if d in domain_sheets]
        ordered += [d for d in domain_sheets if d not in ordered]  # any extras last

        for key in ordered:
            ws = wb[key]
            domain = ParsedDomain(
                key=key,
                name=_clean(ws.cell(1, 1).value) or key,
                scale=_parse_scale(ws),
                **refs.get(key, {}),
            )

            current: ParsedControl | None = None
            for r in range(DATA_START_ROW, ws.max_row + 1):
                col_a = _clean(ws.cell(r, COL_ID).value)
                col_b = _clean(ws.cell(r, COL_ITEM).value)
                if not col_a:
                    continue
                if _is_item(col_a, col_b):
                    if current is None:  # item without a preceding header
                        current = ParsedControl(code="GERAL", name="Geral")
                        domain.controls.append(current)
                    current.questions.append(
                        ParsedQuestion(
                            code=col_a,
                            text=col_b or col_a,
                            guidance=_clean(ws.cell(r, COL_GUIDANCE).value) or None,
                        )
                    )
                else:
                    code, name = _split_header(col_a)
                    current = ParsedControl(code=code, name=name)
                    domain.controls.append(current)

            # Drop empty controls (headers with no items).
            domain.controls = [c for c in domain.controls if c.questions]
            framework.domains.append(domain)

        return framework
